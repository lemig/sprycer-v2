"""Tests for the offer export — the byte-identical I/O contract.

Covers the eng review TODOs:
  - TODO #1: format_euro edge cases including >=€1000 thousands separator
  - TODO #5: CSV byte details (LF only, no BOM, lowercase booleans, empty cells)
  - 1C: reviewed_text 3 cutover-relevant variants
  - dynamic Competitor N columns from MainCompetition.position
  - Cheapest competitors comma-join (single + tied)
"""
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.exporters.offer_export import (
    export_headers,
    render_rows_for_retailer,
    reviewed_text,
    to_csv_bytes,
    to_xlsx_bytes,
)
from core.models import (
    Brand,
    Channel,
    MainCompetition,
    Matching,
    Offer,
    PriceObservation,
    Retailer,
    Review,
    Website,
)
from core.money import format_euro


# ---- format_euro edge cases (TODO #1) -----------------------------------


class TestFormatEuro:
    @pytest.mark.parametrize('cents,expected', [
        (None, ''),
        (0, '€0'),
        (1, '€0.01'),
        (50, '€0.50'),
        (100, '€1'),
        (298, '€2.98'),
        (300, '€3'),         # whole euro -> no decimals (verified in exports.csv)
        (304, '€3.04'),
        (95417, '€954.17'),  # near-max from exports.csv sample
        (123456, '€1,234.56'),  # >=€1000 thousands separator
        (525100, '€5,251'),   # whole >=€1000 thousand
        (97827328, '€978,273.28'),  # very large
    ])
    def test_format(self, cents, expected):
        assert format_euro(cents) == expected


# ---- reviewed_text variants (eng review 1C) -----------------------------


@pytest.mark.django_db
class TestReviewedText:
    @pytest.fixture
    def setup(self):
        sch = Retailer.objects.create(name='Schleiper')
        rp = Retailer.objects.create(name='Rougier & Plé')
        gf = Retailer.objects.create(name='Le Géant des Beaux-Arts (FR)')
        gb = Retailer.objects.create(name='Le Géant des Beaux-Arts (BE)')
        web = Website.objects.create(host='www.schleiper.com')
        ch = Channel.objects.create(name='schleiper.com/eshopexpress', retailer=sch, website=web)
        offer = Offer.objects.create(
            retailer=sch, channel=ch, website=web, sku='X', name='Item', public=True,
        )
        MainCompetition.objects.create(retailer=sch, competitor=rp, position=1)
        MainCompetition.objects.create(retailer=sch, competitor=gf, position=2)
        MainCompetition.objects.create(retailer=sch, competitor=gb, position=3)
        return {'sch': sch, 'rp': rp, 'gf': gf, 'gb': gb, 'offer': offer,
                'main_ids': [rp.id, gf.id, gb.id]}

    def test_no_reviews_yet(self, setup):
        text = reviewed_text(setup['offer'], setup['sch'], setup['main_ids'])
        assert text == 'Competitors offers not yet reviewed'

    def test_partial_reviews(self, setup):
        Review.objects.create(offer=setup['offer'], retailer=setup['sch'], competitor=setup['rp'])
        text = reviewed_text(setup['offer'], setup['sch'], setup['main_ids'])
        assert text == 'Some competitors offers reviewed'

    def test_all_reviewed(self, setup):
        Review.objects.create(offer=setup['offer'], retailer=setup['sch'], competitor=setup['rp'])
        Review.objects.create(offer=setup['offer'], retailer=setup['sch'], competitor=setup['gf'])
        Review.objects.create(offer=setup['offer'], retailer=setup['sch'], competitor=setup['gb'])
        text = reviewed_text(setup['offer'], setup['sch'], setup['main_ids'])
        assert text == 'Main competitors offers reviewed'

    def test_no_main_competitors(self, setup):
        text = reviewed_text(setup['offer'], setup['sch'], [])
        assert text == 'Competitors offers not yet reviewed'


# ---- Dynamic header count (1C: drives Competitor N export columns) ------


@pytest.mark.django_db
class TestDynamicHeaders:
    def _make_retailer(self, name='Schleiper', n_competitors=0):
        r = Retailer.objects.create(name=name)
        for i in range(1, n_competitors + 1):
            comp = Retailer.objects.create(name=f'Competitor{i}')
            MainCompetition.objects.create(retailer=r, competitor=comp, position=i)
        return r

    def test_zero_main_competitors(self):
        r = self._make_retailer(n_competitors=0)
        assert len(export_headers(r)) == 12

    def test_one_main_competitor(self):
        r = self._make_retailer(n_competitors=1)
        assert len(export_headers(r)) == 18

    def test_three_main_competitors(self):
        r = self._make_retailer(n_competitors=3)
        # 12 fixed + 6×3 = 30, matches user-provided exports.csv shape
        assert len(export_headers(r)) == 30

    def test_static_headers_exact_order(self):
        r = self._make_retailer(n_competitors=0)
        h = export_headers(r)
        assert h == [
            'Sprycer ID', 'Channel', 'Retailer', 'SKU', 'Name', 'Price',
            'Price date', 'Public', 'Reviewed', 'Cheapest competitors',
            'Cheapest competitors price', 'Cheapest competitors skus',
        ]

    def test_competitor_block_order(self):
        r = self._make_retailer(n_competitors=1)
        h = export_headers(r)
        assert h[12:] == [
            'Competitor 1', 'Competitor 1 sku', 'Competitor 1 list_price',
            'Competitor 1 price', 'Competitor 1 shipping charges',
            'Competitor 1 price date',
        ]


# ---- CSV byte details (TODO #5 / codex #4) -----------------------------


class TestCsvBytes:
    def test_no_bom_at_start(self):
        out, _ = to_csv_bytes(['A', 'B'], [{'A': 1, 'B': 2}])
        assert not out.startswith(b'\xef\xbb\xbf')
        assert out.startswith(b'A,B\n')

    def test_lf_line_endings_only(self):
        out, _ = to_csv_bytes(['A'], [{'A': 1}, {'A': 2}])
        # Has LFs but no CRs
        assert b'\r' not in out
        assert out.count(b'\n') == 3  # header + 2 rows + trailing newline by writer

    def test_bool_lowercase(self):
        out, _ = to_csv_bytes(['Public'], [{'Public': True}, {'Public': False}])
        text = out.decode('utf-8')
        assert 'Public\ntrue\nfalse\n' == text

    def test_none_becomes_empty_cell(self):
        out, _ = to_csv_bytes(['A', 'B'], [{'A': None, 'B': 'x'}])
        text = out.decode('utf-8')
        assert text == 'A,B\n,x\n'

    def test_count_returned(self):
        _, count = to_csv_bytes(['A'], [{'A': 1}, {'A': 2}, {'A': 3}])
        assert count == 3

    def test_utf8_accented_chars_preserved(self):
        out, _ = to_csv_bytes(['Name'], [{'Name': 'Le Géant des Beaux-Arts'}])
        assert 'Géant'.encode('utf-8') in out

    def test_quoting_minimal_for_comma_in_value(self):
        out, _ = to_csv_bytes(['Name'], [{'Name': 'a, b, c'}])
        # Comma-containing cell gets quoted; others stay unquoted
        assert b'"a, b, c"' in out


# ---- Full integration: render rows for a Schleiper-shaped retailer ------


@pytest.mark.django_db
class TestRenderEndToEnd:
    @pytest.fixture
    def world(self):
        # Schleiper + 3 competitors mirroring production exports.csv
        sch = Retailer.objects.create(name='Schleiper')
        rp = Retailer.objects.create(name='Rougier & Plé')
        gf = Retailer.objects.create(name='Le Géant des Beaux-Arts (FR)')
        gb = Retailer.objects.create(name='Le Géant des Beaux-Arts (BE)')
        for i, comp in enumerate([rp, gf, gb], start=1):
            MainCompetition.objects.create(retailer=sch, competitor=comp, position=i)

        sch_web = Website.objects.create(host='www.schleiper.com')
        rp_web = Website.objects.create(host='www.rougier-ple.fr')
        gf_web = Website.objects.create(host='www.geant-beaux-arts.fr')
        gb_web = Website.objects.create(host='www.geant-beaux-arts.be')

        sch_ch = Channel.objects.create(name='schleiper.com/eshopexpress', retailer=sch, website=sch_web)
        rp_ch = Channel.objects.create(name='rougier-ple.fr', retailer=rp, website=rp_web)
        gf_ch = Channel.objects.create(name='www.geant-beaux-arts.fr', retailer=gf, website=gf_web)
        gb_ch = Channel.objects.create(name='www.geant-beaux-arts.be', retailer=gb, website=gb_web)

        # Schleiper's offer for the marker
        sch_offer = Offer.objects.create(
            id=124397, retailer=sch, channel=sch_ch, website=sch_web,
            sku='WINPMY724', name='Winsor & Newton ProMarker n° 724', public=True,
        )
        PriceObservation.objects.create(offer=sch_offer, price_cents=298)

        # Three competitor offers
        rp_offer = Offer.objects.create(
            retailer=rp, channel=rp_ch, website=rp_web,
            sku='295080', name='ProMarker 724 RP', public=True,
        )
        PriceObservation.objects.create(offer=rp_offer, price_cents=300)  # €3 — cheapest
        gf_offer = Offer.objects.create(
            retailer=gf, channel=gf_ch, website=gf_web,
            sku='35940447', name='ProMarker 724 GF', public=True,
        )
        PriceObservation.objects.create(offer=gf_offer, price_cents=304)
        gb_offer = Offer.objects.create(
            retailer=gb, channel=gb_ch, website=gb_web,
            sku='35940447', name='ProMarker 724 GB', public=True,
        )
        PriceObservation.objects.create(offer=gb_offer, price_cents=302)

        # Confirmed matchings: Schleiper -> each competitor
        for c_offer in (rp_offer, gf_offer, gb_offer):
            Matching.objects.create(
                offer=sch_offer, competing_offer=c_offer,
                status=Matching.Status.CONFIRMED, source=Matching.Source.LEGACY_IMPORTED,
            )

        # Mark all competitors reviewed by Schleiper
        for c in (rp, gf, gb):
            Review.objects.create(offer=sch_offer, retailer=sch, competitor=c)

        return {'sch': sch, 'rp': rp, 'gf': gf, 'gb': gb, 'offer': sch_offer}

    def test_first_row_matches_legacy_shape(self, world):
        headers, row_iter = render_rows_for_retailer(world['sch'])
        rows = list(row_iter)
        assert len(rows) == 1
        row = rows[0]

        assert row['Sprycer ID'] == 124397
        assert row['Channel'] == 'schleiper.com/eshopexpress'
        assert row['Retailer'] == 'Schleiper'
        assert row['SKU'] == 'WINPMY724'
        assert row['Price'] == '€2.98'
        assert row['Public'] is True
        assert row['Reviewed'] == 'Main competitors offers reviewed'
        assert row['Cheapest competitors'] == 'Rougier & Plé'
        assert row['Cheapest competitors price'] == '€3'
        assert row['Cheapest competitors skus'] == '295080'

        # Competitor 1 = Rougier & Plé (position 1)
        assert row['Competitor 1'] == 'Rougier & Plé'
        assert row['Competitor 1 sku'] == '295080'
        assert row['Competitor 1 price'] == '€3'

        # Competitor 2 = Le Géant FR (position 2)
        assert row['Competitor 2'] == 'Le Géant des Beaux-Arts (FR)'
        assert row['Competitor 2 price'] == '€3.04'

        # Competitor 3 = Le Géant BE (position 3)
        assert row['Competitor 3'] == 'Le Géant des Beaux-Arts (BE)'
        assert row['Competitor 3 price'] == '€3.02'

    def test_to_csv_full_row_string_match(self, world):
        headers, rows = render_rows_for_retailer(world['sch'])
        out, count = to_csv_bytes(headers, rows)
        text = out.decode('utf-8')
        lines = text.split('\n')
        # 1 header + 1 row + trailing newline
        assert count == 1
        assert len(lines) == 3 and lines[2] == ''
        # Header order
        assert lines[0].split(',')[:6] == ['Sprycer ID', 'Channel', 'Retailer', 'SKU', 'Name', 'Price']
        # Row 1 starts with the offer's id and channel
        assert lines[1].startswith('124397,schleiper.com/eshopexpress,Schleiper,WINPMY724,')
        # Money column shows the no-decimal-on-whole-euro behavior
        assert ',€2.98,' in lines[1]
        assert ',€3,' in lines[1]
        # Boolean is lowercase
        assert ',true,' in lines[1]
        # Reviewed cell text exact
        assert 'Main competitors offers reviewed' in lines[1]


@pytest.mark.django_db
class TestXlsxRoundtrip:
    def test_xlsx_cells_match_row_dict(self):
        sch = Retailer.objects.create(name='Schleiper')
        web = Website.objects.create(host='www.schleiper.com')
        ch = Channel.objects.create(name='schleiper.com/eshopexpress', retailer=sch, website=web)
        offer = Offer.objects.create(
            retailer=sch, channel=ch, website=web, sku='X1', name='Item 1', public=True,
        )
        PriceObservation.objects.create(offer=offer, price_cents=298)

        # competing_offers_only=False so the default Schleiper filter (only
        # offers with confirmed matchings) doesn't drop the lone test offer.
        headers, row_iter = render_rows_for_retailer(sch, competing_offers_only=False)
        out, count = to_xlsx_bytes(headers, row_iter)
        assert count == 1

        wb = load_workbook(BytesIO(out))
        ws = wb.active
        cell_headers = [c.value for c in ws[1]]
        assert cell_headers[:6] == ['Sprycer ID', 'Channel', 'Retailer', 'SKU', 'Name', 'Price']
        cell_row = [c.value for c in ws[2]]
        # SKU + name appear in the right cells
        assert 'X1' in cell_row
        assert 'Item 1' in cell_row
        assert '€2.98' in cell_row
